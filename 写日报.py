# -*- coding: utf-8 -*-
from lxml import etree
import requests
import  xlsxwriter
import datetime
from openpyxl import load_workbook
import re
# 网站888文章更新函数

def drinks888(url):
    datas=[]        
    html=requ_html(url)
    html2=requ_html("http://www.drinks888.com/news/53/2.html")     
    url=url.split('/news')

    # 第一页的三条数据
    title=html.xpath('//div[@class="news_main2"]/dl[position()>1]/dd/h3/a/text()')
    urls=html.xpath('//div[@class="news_main2"]/dl[position()>1]/dd/h3/a/@href')
    for i in range(0,3):
        datas.append(url[0]+urls[i])
        datas.append(title[i].strip())
        # 第二页的两条
    title2=html2.xpath('//div[@class="news_main2"]/dl/dd/h3/a/text()')
    urls2=html.xpath('//div[@class="news_main2"]/dl/dd/h3/a/@href')
    for i in range(0,2):
        datas.append(url[0]+urls2[i])
        datas.append(title2[i].strip())
    wb = load_workbook(r'C:\Users\Administrator\Desktop\唐富\唐富\5月工作表\各类长尾词统计.xlsx')
    wbs=wb['所有文章更新链接']
    n=wbs.max_row+2
    for i in range(0,10,2):
        wbs['D{}'.format(n)] = datas[i]
        wbs['E{}'.format(n)] = datas[i+1]
        n=n+1
    wb.save(r'C:\Users\Administrator\Desktop\唐富\唐富\5月工作表\各类长尾词统计.xlsx')   
    print('链接已写入各类关键词表中')     
    return datas

# 网站999文章更新的函数
def drinks999(url):
    datas=[]
    html=requ_html(url)
    url=url.split('/news')
    list_p=html.xpath('//dl[@class="dl_220"]/dd/a/text()')
    list_url=html.xpath('//dl[@class="dl_220"]/dd/a/@href')
    for i in range(0,5):
        title_text=list_p[i].encode('ISO-8859-1').decode('UTF-8')
        url_href=url[0]+list_url[i]
        # worksheet.write(9+i,2,url_href)
        # worksheet.write(9+i,3,title_text)
        datas.append(url_href)
        datas.append(title_text.strip())
    return datas


# 网站aaa的更新函数
def drinksaaa(url):
    datas=[]
    html=requ_html(url)
    url=url.split('/a')
    list_p=html.xpath('//ul[@class="fc_pro"]/li/a/h3/text()')[0].encode('ISO-8859-1').decode('UTF-8')
    list_url=url[0]+html.xpath('//ul[@class="fc_pro"]/li/a/@href')[0]
    # url_href=url[0]+list_url
    # worksheet.write(15,2,list_p)
    # worksheet.write(15,3,url_href)
    # print(list_p)
    # print(url_href)
    
    datas.append(list_url)
    datas.append(list_p.strip())

    return datas


# 网站bbb、ccc、ddd、eee、fff、kkk/mmm文章更新函数
def drinksbcdef(url):
    # header={
    #     'User-Agent':'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/55.0.2883.87 Safari/537.36',
    #     'Cookie':'bdshare_firstime=1574388395758; yjs_id=aHR0cDovL3d3dy5kcmlua3NiYmIuY29tL2p3dGNiYmIwMjgvfDE1ODc2MjMzNjkyNjU; DedeUserID=1; DedeUserID__ckMd5=9880aa35c1f9b840; DedeLoginTime=1597627802; DedeLoginTime__ckMd5=0ae3c362e99a6e1f; Hm_lvt_8b1432e8c92af2a6596d7f512dabf0f0=1597281289,1597367738,1597628086,1597657062; Hm_lpvt_8b1432e8c92af2a6596d7f512dabf0f0=1597657062'
    # }
    datas=[]
    html=requ_html(url)
    # print(html)    
    list_p=html.xpath('//div[@class="txt"]/h2/a/text()')[0].encode('ISO-8859-1').decode('UTF-8')
    url=url.split('/a')
    list_url=url[0]+html.xpath('//div[@class="txt"]/h2/a/@href')[0]       
    
    datas.append(list_url)
    datas.append(list_p.strip())
    return datas    


# 网站qqq文章更新函数  
def drinksqqq(url):
    datas=[]
    html=requ_html(url)
    url=url.split('/a')
    list_title=html.xpath("//div[@class='news_con']/dl/dt/a/text()")[0].encode('ISO-8859-1').decode('UTF-8')
    list_url=url[0]+html.xpath("//div[@class='news_con']/dl/dt/a/@href")[0]
    datas.append(list_url)
    datas.append(list_title.strip())
    # print(list_url)
    # print(list_title.strip())
    return datas


# 网站rrr更新函数
def drinksrrr(url):
    datas=[]
    html=requ_html(url)
    url=url.split('/a')
    list_title=html.xpath('//ul[@class="cpshow"]/li/h4/a/text()')[0].encode('ISO-8859-1').decode('UTF-8')
    list_url=url[0]+html.xpath('//ul[@class="cpshow"]/li/h4/a/@href')[0]
    # print(list_title)
    # print(list_url)
    datas.append(list_url)
    datas.append(list_title.strip())
    return datas
# 网站sss、ggg更新函数==============


def drinkssss(url):
    datas=[]    
    html=requ_html(url)
    url=url.split('/a')
    list_title=html.xpath('//div[@class="agent_con"]/dl/dd/h4/a/text()')[0].encode('ISO-8859-1').decode('UTF-8')
    list_url=url[0]+html.xpath('//div[@class="agent_con"]/dl/dd/h4/a/@href')[0]
    datas.append(list_url)
    datas.append(list_title.strip())
    return datas


# 网站hhh/nnn更新函数======================
def drinkshhh(url):
    datas=[]
    html=requ_html(url)
    url=url.split('/a')
    list_title=html.xpath('//div[@class="VieList"]/h3/a/text()')[0].encode('ISO-8859-1').decode('UTF-8')
    list_url=url[0]+html.xpath('//div[@class="VieList"]/h3/a/@href')[0]
    datas.append(list_url)
    datas.append(list_title.strip())
    return datas


# 网站jjj更新函数======================
def drinksjjj(url):
    datas=[]
    html=requ_html(url)
    url=url.split('/a')
    list_title=html.xpath('//div[@class="pro_main"]/dl/dd/a/text()')[0].encode('ISO-8859-1').decode('UTF-8')
    list_url=url[0]+html.xpath('//div[@class="pro_main"]/dl/dd/a/@href')[0]
    datas.append(list_url)
    datas.append(list_title.strip())
    return datas


# 网站lll函数=================
def drinkslll(url):
    datas=[]
    html=requ_html(url)
    url=url.split('/a')    
    list_title=html.xpath('//div[@class="agent_con"]/dl/dd/a/text()')[0].encode('ISO-8859-1').decode('UTF-8')
    list_url=url[0]+html.xpath('//div[@class="agent_con"]/dl/dd/a/@href')[0]
    datas.append(list_url)
    datas.append(list_title.strip())
    return datas


# 网站ooo更新函数=================
def drinksooo(url):
    datas=[]
    
    html=requ_html(url)
    url=url.split('/a') 
    list_title=html.xpath("//div[@class='news_con']/dl/dd/span/a/text()")[0].encode('ISO-8859-1').decode('UTF-8')
    list_url=url[0]+html.xpath("//div[@class='news_con']/dl/dd/span/a/@href")[0]
    datas.append(list_url)
    datas.append(list_title.strip())
    return datas


# 网站ppp函数=====================
def drinksppp(url):
    html=requ_html(url)
    datas=[]
    url=url.split('/a') 
    list_title=html.xpath("//dl[@class='pd_list_dl']/dd/a/text()")[0].encode('ISO-8859-1').decode('UTF-8')
    list_url=url[0]+html.xpath("//dl[@class='pd_list_dl']/dd/a/@href")[0]
    datas.append(list_url)
    datas.append(list_title.strip())
    return datas

# 通用===================================
def requ_html(url):
    headers={
    'user-agent':'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/55.0.2883.87 Safari/537.36'}
    res=requests.get(url,headers=headers)
    html=etree.HTML(res.text)
    return html


# 搜狐函数====================
def souhu(url):
    datas=[]
    html=requ_html(url)
    # 第一个搜狐号
    title=html.xpath('//ul/li[position()<6]/article/div/h4/a/text()')
    urls=html.xpath('//ul/li[position()<6]/article/div/h4/a/@href')    
    for i in range(0,5):
        datas.append('https:'+urls[i])
        datas.append(title[i].strip())
    # 第二个搜狐号
    html2=requ_html('https://mp.sohu.com/profile?xpt=NTU0MmJiMDgtYzMxNC00MzRjLWIxOWUtNTdmMzk1MDBhZDg1&_f=index_pagemp_2&spm=smpc.content.author.3.159825042146246R1zHx')
    title2=html2.xpath('//ul/li[position()<6]/article/div/h4/a/text()')
    urls2=html2.xpath('//ul/li[position()<6]/article/div/h4/a/@href')
    for i in range(0,5):
        datas.append('https:'+urls2[i])
        datas.append(title2[i].strip())   
    return datas


# 博客======================================
def boke(url):
    datas=[]
    html=requ_html(url)
    title=html.xpath('//div[@class="articleList"]/div[position()<6]/p/span/a/text()')
    urls=html.xpath('//div[@class="articleList"]/div[position()<6]/p/span/a/@href')
    for i in range(0,5):
        
        datas.append(urls[i])
        datas.append(title[i].encode('ISO-8859-1').decode('UTF-8'))
    return datas


# 创头条===============================
def chuangtout(url):
    datas=[]
    html=requ_html(url)      
    urls=html.xpath('//h2/a/@href')    
    title=html.xpath('//h2/a/text()')
    for i in range(0,5):    
        datas.append("http://www.ctoutiao.com"+urls[i].replace('\\','').replace('"',''))
        n=title[i].replace(r'\r\n','').encode('utf-8').decode('unicode_escape').strip()
        datas.append(n.strip())
    return datas


# 818同城发布==============================
def get_818(url):
    datas=[]
    headers={
    'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/73.0.3683.86 Safari/537.36'
            }
   
    # s = requests.session() #建立一个Session
    form_data = {   
        '__VIEWSTATE': '/wEPDwUKLTk4NDU4OTMzMWRk2n6a8smWQwShiX7p3Dw0lcMKWtXllwxr+A6O7EuLWfE=',
        '__VIEWSTATEGENERATOR': 'D399C246',
        'username': 'a463459227',    
        'userpassword':'463459227',
        'Button1': '登录',
    }

    session = requests.session()
    response = session.post(url,headers=headers,data=form_data) #session登录网站
    url="http://www.818u.com/members/infolist.aspx?uid=447889"
    response = session.get(url,headers=headers) #session浏览页面
    html=etree.HTML(response.text)
    title=html.xpath('//tr[position()<7]/td/a[@class="f0"]/text()')
    urls=html.xpath('//tr[position()<7]/td/a[@class="f0"]/@href')
    for i in range(0,5):
        datas.append('http://www.818u.com'+urls[i])
        datas.append(title[i])
    return datas

# 百姓网==============================
def get_baixing():
    datas=[]
    url='https://www.baixing.com/u/92147482/?src=vad_listing_7'
    headers={
    'user-agent':'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/55.0.2883.87 Safari/537.36'}
    res=requests.get(url,headers=headers)    
    html=etree.HTML(res.text)    
    title=html.xpath('//ul[@class="list-ad-items"]/li[position()<6]/div/div[1]/a/text()')
    baixing_url=html.xpath('//ul[@class="list-ad-items"]/li[position()<6]/div/div[1]/a/@href')
    for i in range(0,5):
        datas.append(baixing_url[i].replace('?from=',''))
        datas.append(title[i])
    return datas 
# 知乎
def zhihu():
    datas=[]
    url='https://www.zhihu.com/people/ding-ni-ge-fei-68-18/posts'
    headers={
    'user-agent':'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/55.0.2883.87 Safari/537.36'
    }
    res=requests.get(url,headers=headers)
    html=res.text
    url_list=re.search('<script id="js-initialData" type="text/json">(.*?)</script>',html)[0]
    # title=html.xpath('//h2[@class="ContentItem-title"][position()<6]/a/text()')
    urls=re.findall('"author".*?},"url":"(.*?)","commentPermission"',url_list)
    title_list=re.findall('"linkbox":{.*?},"title":"(.*?)","voting',url_list)

    # print(t[0].encode('utf8').decode('unicode_escape'))
    for i in range(5,10):
        datas.append(urls[i].encode('utf8').decode('unicode_escape'))
        datas.append(title_list[i])
    return datas       

#b站栏目文章抓取
def bili_cont():
    datas=[]
    url='https://api.bilibili.com/x/space/article?mid=702710400&pn=1&ps=12&sort=publish_time&jsonp=jsonp'
    res=requests.get(url)
    html=res.text
    title=re.findall(r'日常.*?"title":"(\S+?)"',html)
    url_id=re.findall(r'{"id":(\d+),"category',html)
    for i in range(0,5):
        datas.append('https://www.bilibili.com/read/cv{}'.format(url_id[i]))
        datas.append(title[i])
    return datas
# 主函数===================
def main():  
    now_time =datetime.datetime.now().strftime('%Y-%m-%d')
    workbook = xlsxwriter.Workbook("C:/Users/Administrator/Desktop/唐富/唐富/5月工作表/{}工作日报.xlsx".format(now_time)) # 创建excel
    worksheet = workbook.add_worksheet("first_sheet") # 创建sheet
    merge_format = workbook.add_format({
    'bold': True,
    "align": "center",  # 对齐方式
    "valign": "vcenter",  # 字体对齐方式
    "font_color": "red"  })
    style=workbook.add_format({
    "font_color": "red"     })
    worksheet.merge_range('B3:C3','网站文章更新',merge_format)
    # worksheet.merge_range('B3:C4', '网站文章更新', merge_format)
    # worksheet.write(2, 1, '网站文章更新')
    worksheet.write(3, 1, '网站888文章更新',style)
    worksheet.write(3, 3, '818同城发布',style)
    worksheet.write(9, 1, '网站999文章更新',style)
    worksheet.write(9, 3, '创头条',style)
    # 设置宽度
    worksheet.set_column(1,4, 45)
    # worksheet.write(10, 2,'url')
    # worksheet.write(10, 3,'title')
    worksheet.write(15, 1, '其他网站文章更新',style)
    worksheet.write(34, 1, '搜狐新浪',style)
    # worksheet.write(16, 1,'url')
    # worksheet.write(16, 2,'title')
    # 网站888文章
    print('网站888文章开始')
    datas=drinks888("http://www.drinks888.com/news/53/1.html")
    # 激活 worksheet
    n=4
    for i in range(0,10,2):
        worksheet.write(n,1,datas[i])
        worksheet.write(n,2,datas[i+1])
        n=n+1
   
    print('网站999文章开始')
    datas=drinks999('http://www.drinks999.com/news')
    n=10  
    for i in range(0,10,2):   
        worksheet.write(n,1,datas[i])
        worksheet.write(n,2,datas[i+1])
        n=n+1   
    print('网站aaa文章开始')
    datas_aaa=drinksaaa("http://www.drinksaaa.com/a/xinwenzixun")
    worksheet.write(16,1,datas_aaa[0])
    worksheet.write(16,2,datas_aaa[1])
    print('网站bbb更新开始')
    datas_bbb=drinksbcdef("http://www.drinksbbb.com/a/xinwenzixun/")
    worksheet.write(17,1,datas_bbb[0])
    worksheet.write(17,2,datas_bbb[1])
    print('网站ccc更新开始')
    datas_ccc=drinksbcdef("http://www.drinksccc.com/a/xingyezixun")
    worksheet.write(18,1,datas_ccc[0])
    worksheet.write(18,2,datas_ccc[1])
    print('网站ddd更新开始')
    datas_ddd=drinksbcdef("http://www.drinksddd.com/a/xingyezixun")
    worksheet.write(19,1,datas_ddd[0])
    worksheet.write(19,2,datas_ddd[1])
    print('网站eee更新开始')
    datas_eee=drinksbcdef("http://www.drinkseee.com/a/xinwenzixun")
    worksheet.write(20,1,datas_eee[0])
    worksheet.write(20,2,datas_eee[1])
    print('网站fff更新开始')
    datas_fff=drinksbcdef("http://www.drinksfff.com/a/xingyezixun")
    worksheet.write(21,1,datas_fff[0])
    worksheet.write(21,2,datas_fff[1])
    print('网站ggg更新开始')
    datas_ggg=drinkslll("http://www.drinksggg.com/a/xinwenzixun")
    worksheet.write(22,1,datas_ggg[0])
    worksheet.write(22,2,datas_ggg[1])
    print('网站hhh更新开始')
    datas_hhh=drinkshhh("http://www.drinkshhh.com/a/xinwenzixun")
    worksheet.write(23,1,datas_hhh[0])
    worksheet.write(23,2,datas_hhh[1])
    print('网站jjj更新开始')
    datas_jjj=drinksjjj("http://www.drinksjjj.com/news")
    worksheet.write(24,1,datas_jjj[0])
    worksheet.write(24,2,datas_jjj[1])
    print('网站kkk更新开始')
    datas_kkk=drinksbcdef("http://www.drinkskkk.com/a/xinwenzixun")
    worksheet.write(25,1,datas_kkk[0])
    worksheet.write(25,2,datas_kkk[1])
    print('网站mmm更新开始')
    datas_mmm=drinksbcdef("http://www.drinksmmm.com/a/xinwenzixun")
    worksheet.write(26,1,datas_mmm[0])
    worksheet.write(26,2,datas_mmm[1])
    print('网站nnn更新开始')
    datas_nnn=drinkshhh("http://www.drinksnnn.com/a/xinwenzixun")
    worksheet.write(27,1,datas_nnn[0])
    worksheet.write(27,2,datas_nnn[1])
    print('网站ooo更新开始')
    datas_ooo=drinksooo("http://www.drinksooo.com/a/xinwenzixun")
    worksheet.write(28,1,datas_ooo[0])
    worksheet.write(28,2,datas_ooo[1])
    print('网站ppp更新开始')
    datas_ppp=drinksppp("http://www.drinksppp.com/a/xinwenzixun")
    worksheet.write(29,1,datas_ppp[0])
    worksheet.write(29,2,datas_ppp[1])
    print('网站qqq更新开始')
    datas_qqq=drinksqqq("http://www.drinksqqq.com/news")
    worksheet.write(30,1,datas_qqq[0])
    worksheet.write(30,2,datas_qqq[1])
    print('网站rrr更新开始')
    datas_rrr=drinksrrr("http://www.drinksrrr.com/a/xinwenzixun")
    worksheet.write(31,1,datas_rrr[0])
    worksheet.write(31,2,datas_rrr[1])
    print('网站sss更新开始')
    datas_sss=drinkssss("http://www.drinkssss.com/a/xinwenzhongxin")
    worksheet.write(32,1,datas_sss[0])
    worksheet.write(32,2,datas_sss[1])
    print('网站lll更新开始')
    datas_lll=drinkslll("http://www.drinkslll.com/a/xinwenzixun")
    worksheet.write(33,1,datas_lll[0])
    worksheet.write(33,2,datas_lll[1])

    print('搜狐开始')
    datas=souhu("https://mp.sohu.com/profile?xpt=Y2hvdWZlaTkwN0Bzb2h1LmNvbQ==&_f=index_pagemp_2&spm=smpc.content.author.3.1598248792443G5ANNWn")
    n=35  
    for i in range(0,20,2):   
        worksheet.write(n,1,datas[i])
        worksheet.write(n,2,datas[i+1])
        n=n+1
    print('博客开始')
    # 博客
    # datas=boke("http://blog.sina.com.cn/s/articlelist_5226711056_0_1.html")
    datas=bili_cont()
    n=45
    for i in range(0,8,2):
        worksheet.write(n,1,datas[i])
        worksheet.write(n,2,datas[i+1])  
        n=n+1
    # 创头条
    print('创头条开始')
    datas=chuangtout("http://www.ctoutiao.com/ajax_new/ajax_data.php?page=newCompany&act=getPosts&uid=1729092&type=getPosts&pageno=1")
    n=10
    for i in range(0,10,2):
        worksheet.write(n,3,datas[i])
        worksheet.write(n,4,datas[i+1])  
        n=n+1

    # 818同城发布
    print('818同城开始')
    datas=get_818('http://www.818u.com/members/login.aspx')
    # datas=get_baixing()
    # datas=zhihu()
    # datas=bili_cont()
    n=4
    for i in range(0,10,2):
        worksheet.write(n,3,datas[i])
        worksheet.write(n,4,datas[i+1])  
        n=n+1 
    workbook.close()  # 关闭excel写入
    print('日报已经写完了')

main()
    


    
        