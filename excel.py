
from openpyxl import load_workbook
# 实例化

# 激活 worksheet
wb = load_workbook(r'C:\Users\Administrator\Desktop\唐富\唐富\5月工作表\各类长尾词统计.xlsx')
wbs=wb['所有文章更新链接']

n=wbs.max_row+2

for i in range(0,10,2):
    wbs['D{}'.format(n)] = datas[i]
    wbs['E{}'.format(n)] = datas[i+1]
    n=n+1
    
# print(wbs.max_row)
# wb.save(r'C:\Users\Administrator\Desktop\唐富\唐富\5月工作表\各类长尾词统计.xlsx')
